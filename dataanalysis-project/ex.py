import openpyxl
import random
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "人员安排"
title = ('周次','值日人员')
print(type(title))
for col_index,title in enumerate(title):
    sheet.cell(row=1, column=col_index+1, value=title)
wb.save('打扫卫生安排.xlsx')
k='''代欣江
樊国梁
苟冰灵
唐凡琪
贾晨晗
谢永桂
毕可诺
郭宇航
李应鑫
杨玉洁
刘渝川
毛金磊
夏雨婷
何映鹏
周鑫
何朋飞'''
m=("八","九","十","十一","十二","十三","十四","十五")
names = k.split('\n')
random.shuffle(names)
for i in range(0,len(names),2):
    name1=names[i]
    name2=names[i+1] if (i+1)<len(names) else ""
    group_name=f"{name1}，{name2}"
    sheet.cell(row=i//2+2,column=2,value=group_name)
    sheet.cell(row=i//2+2,column=1,value=m[i//2])
wb.save('打扫卫生安排.xlsx')